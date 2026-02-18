"""
Универсальная библиотека для создания Excel-файлов из спецификаций
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class SpecificationExcelBuilder:
    """Класс для создания Excel-файлов из данных спецификаций"""
    
    def __init__(self, output_path):
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
    def create_sheet(self, sheet_name, data, column_widths=None):
        """
        Создает лист с данными
        
        Args:
            sheet_name: название листа
            data: список списков (строки таблицы)
            column_widths: словарь {индекс_колонки: ширина} или None для дефолтных значений
        """
        ws = self.wb.create_sheet(sheet_name)
        
        # Стили
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell_font = Font(name='Arial', size=9)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заполнение данными
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = cell_font
        
        # Ширина колонок
        if column_widths is None:
            # Дефолтные ширины для 9-колоночной спецификации
            column_widths = {
                'A': 8, 'B': 50, 'C': 30, 'D': 15, 
                'E': 20, 'F': 12, 'G': 10, 'H': 12, 'I': 25
            }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        return ws
    
    def save(self):
        """Сохраняет файл"""
        self.wb.save(self.output_path)
        

class SpecificationDataExtractor:
    """Класс для извлечения и структурирования данных из спецификаций"""
    
    @staticmethod
    def create_header(use_code_column=True):
        """
        Создает стандартный заголовок таблицы
        
        Args:
            use_code_column: если True, использует "Код продукции", иначе "Код оборудования"
        """
        code_name = "Код продукции" if use_code_column else "Код оборудования, изделия, материала"
        supplier_name = "Поставщик" if use_code_column else "Завод - изготовитель"
        
        return [
            "Позиция" if use_code_column else "Позиция",
            "Наименование и техническая характеристика",
            "Тип, марка, обозначение документа, опросного листа",
            code_name,
            supplier_name,
            "Единица измерения",
            "Количество" if use_code_column else "Количество",
            "Масса единицы, кг" if not use_code_column else "Масса 1 ед., кг",
            "Примечание"
        ]
    
    @staticmethod
    def create_row(pos="", name="", type_mark="", code="", supplier="", 
                   unit="", qty="", mass="", note=""):
        """Создает строку таблицы с данными"""
        return [pos, name, type_mark, code, supplier, unit, qty, mass, note]
    
    @staticmethod
    def create_section_header(section_name):
        """Создает заголовок раздела (позиция с названием системы)"""
        return ["", section_name, "", "", "", "", "", "", ""]


# Пример использования
def example_usage():
    """Демонстрация использования библиотеки"""
    
    # Создаем построитель
    builder = SpecificationExcelBuilder('/home/claude/example_spec.xlsx')
    
    # Создаем экстрактор данных
    extractor = SpecificationDataExtractor()
    
    # Формируем данные для листа
    data = [
        extractor.create_header(use_code_column=False),
        extractor.create_row(
            pos="В1",
            name="Хозяйственно-питьевой водопровод",
            type_mark="",
            code="",
            supplier="",
            unit="",
            qty="",
            mass="",
            note=""
        ),
        extractor.create_row(
            pos="",
            name="Трубы стальные электросварные 108х2,8",
            type_mark="ГОСТ 10704-91",
            code="",
            supplier="",
            unit="м",
            qty="12,0",
            mass="7,26",
            note=""
        ),
    ]
    
    # Создаем лист
    builder.create_sheet("Пример", data)
    
    # Сохраняем
    builder.save()
    print("Пример создан!")


if __name__ == "__main__":
    example_usage()
