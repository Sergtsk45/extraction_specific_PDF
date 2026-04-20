"""
@file: test_po_builder.py
@description: Тесты для модуля po_builder (построитель XLSX заказов закупки Odoo).
@location: services/invoice-extractor/backend/tests/test_po_builder.py
"""

from __future__ import annotations
import io
import pytest
import openpyxl

from app.po_builder import build_po_xlsx, PO_HEADERS


# ── Фикстуры ──────────────────────────────────────────────────────────────────

@pytest.fixture
def po_xlsx_path(tmp_path):
    return str(tmp_path / "test_po.xlsx")


@pytest.fixture
def sample_invoice():
    return {
        "invoice_number": "СЧ-2026-042",
        "date": "2026-04-20",
        "supplier": {"inn": "2801234567", "name": "ИП Татаринов В.В."},
        "items": [
            {
                "line_no": 1,
                "name": "Фильтр магнитный фланцевый 100 Ру16",
                "article": "ФМФ-100",
                "unit": "шт",
                "qty": 2,
                "price": 4500.00,
                "vat_rate": 20,
                "vat_amount": 1800.00,
                "amount_w_vat": 10800.00,
            },
            {
                "line_no": 2,
                "name": "Фланец LD плоский кованый 100-10",
                "article": "ФЛД-100",
                "unit": "шт",
                "qty": 4,
                "price": 350.00,
                "vat_rate": 20,
                "vat_amount": 280.00,
                "amount_w_vat": 1680.00,
            },
        ],
    }


# ── Unit-тесты: файл и структура ──────────────────────────────────────────────

def test_creates_file(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    import os
    assert os.path.exists(po_xlsx_path)


def test_sheet_name(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    wb = openpyxl.load_workbook(po_xlsx_path)
    assert "purchase.order" in wb.sheetnames


def test_headers(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    actual = [ws.cell(1, c).value for c in range(1, len(PO_HEADERS) + 1)]
    assert actual == PO_HEADERS


def test_row_count(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    # Заголовок + 2 строки товаров
    assert ws.max_row == 3


def test_empty_items_only_header(po_xlsx_path):
    build_po_xlsx({"invoice_number": "X", "items": []}, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.max_row == 1


# ── Unit-тесты: External ID ────────────────────────────────────────────────────

def test_order_ext_id_format(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    # Обе строки должны иметь одинаковый order id
    assert ws.cell(2, 1).value == "po_2801234567_СЧ-2026-042"
    assert ws.cell(3, 1).value == "po_2801234567_СЧ-2026-042"


def test_line_ext_id_format(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 5).value == "po_2801234567_СЧ-2026-042_line_1"
    assert ws.cell(3, 5).value == "po_2801234567_СЧ-2026-042_line_2"


def test_inn_fallback_when_missing(po_xlsx_path):
    data = {"invoice_number": "Z", "items": [{"line_no": 1, "name": "Товар"}]}
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 1).value == "po_x_Z"


# ── Unit-тесты: заголовок заказа (только первая строка) ──────────────────────

def test_order_header_only_in_first_row(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    # Строка 2: name, partner_id, date_order заполнены
    assert ws.cell(2, 2).value == "СЧ-2026-042"
    assert ws.cell(2, 3).value == "ИП Татаринов В.В."
    assert ws.cell(2, 4).value == "2026-04-20"
    # Строка 3: name, partner_id, date_order пустые (openpyxl даёт None для пустой ячейки)
    assert ws.cell(3, 2).value in (None, "")
    assert ws.cell(3, 3).value in (None, "")
    assert ws.cell(3, 4).value in (None, "")


# ── Unit-тесты: поля строки ────────────────────────────────────────────────────

def test_product_name(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 6).value == "Фильтр магнитный фланцевый 100 Ру16"


def test_product_qty(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 8).value == 2


def test_price_unit_without_vat(po_xlsx_path, sample_invoice):
    """Цена должна быть без НДС (поле price, а не amount_w_vat/qty)."""
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 9).value == 4500.00


def test_unit_of_measure(po_xlsx_path, sample_invoice):
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 10).value == "шт"


def test_tax_format(po_xlsx_path, sample_invoice):
    """НДС 20 → '20%'."""
    build_po_xlsx(sample_invoice, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 11).value == "20%"


def test_no_vat_rate_gives_empty_tax(po_xlsx_path):
    data = {
        "invoice_number": "X",
        "items": [{"line_no": 1, "name": "Товар", "qty": 1, "price": 100}],
    }
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 11).value in (None, "")


def test_vat_rate_empty_string_no_crash(po_xlsx_path):
    """Пустая строка vat_rate раньше давала float('') → 500."""
    data = {
        "invoice_number": "X",
        "items": [{"line_no": 1, "name": "Товар", "qty": 1, "price": 100, "vat_rate": ""}],
    }
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 11).value in (None, "")


def test_vat_rate_string_with_percent(po_xlsx_path):
    data = {
        "invoice_number": "X",
        "items": [{"line_no": 1, "name": "Товар", "qty": 1, "price": 100, "vat_rate": "20%"}],
    }
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 11).value == "20%"


def test_qty_price_as_strings(po_xlsx_path):
    data = {
        "invoice_number": "X",
        "items": [{"line_no": 1, "name": "Товар", "qty": "10", "price": "100,50"}],
    }
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.cell(2, 8).value == 10.0
    assert ws.cell(2, 9).value == 100.5


def test_zero_qty_no_crash(po_xlsx_path):
    data = {
        "invoice_number": "X",
        "items": [{"line_no": 1, "name": "Товар", "qty": 0, "price": 100}],
    }
    build_po_xlsx(data, po_xlsx_path)  # не должно кидать ZeroDivisionError


def test_missing_optional_fields_no_crash(po_xlsx_path):
    """Нет unit, vat_rate, article — файл строится без ошибок."""
    data = {"invoice_number": "X", "items": [{"name": "Минимум"}]}
    build_po_xlsx(data, po_xlsx_path)
    ws = openpyxl.load_workbook(po_xlsx_path)["purchase.order"]
    assert ws.max_row == 2


# ── Integration-тест: endpoint ────────────────────────────────────────────────

MINIMAL_PDF = b"%PDF-1.4 1 0 obj<</Type /Catalog>>endobj\nxref\n0 0\ntrailer<</Size 1>>\nstartxref\n0\n%%EOF"


def test_convert_odoo_po_xlsx_returns_200(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    resp = client.post(
        "/convert",
        data={"file": (io.BytesIO(MINIMAL_PDF), "inv.pdf"), "output": "odoo_po_xlsx"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200


def test_convert_odoo_po_xlsx_content_type(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    resp = client.post(
        "/convert",
        data={"file": (io.BytesIO(MINIMAL_PDF), "inv.pdf"), "output": "odoo_po_xlsx"},
        content_type="multipart/form-data",
    )
    assert "spreadsheetml" in resp.content_type


def test_convert_odoo_po_xlsx_output_mode_header(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    resp = client.post(
        "/convert",
        data={"file": (io.BytesIO(MINIMAL_PDF), "inv.pdf"), "output": "odoo_po_xlsx"},
        content_type="multipart/form-data",
    )
    assert resp.headers.get("X-Output-Mode") == "odoo_po_xlsx"


def _sample() -> dict:
    return {
        "invoice_number": "TEST-PO-001",
        "date": "2026-04-20",
        "supplier": {"inn": "1234567890", "name": "ООО Тест", "kpp": "", "address": "", "bank": {}},
        "buyer": {"inn": "", "name": "", "kpp": "", "address": ""},
        "items": [
            {
                "line_no": 1, "name": "Товар 1", "article": "A1", "unit": "шт",
                "qty": 5, "price": 100.0, "vat_rate": 20,
                "vat_amount": 100.0, "amount_w_vat": 600.0,
            },
        ],
        "totals": {"total_wo_vat": 500.0, "vat_total": 100.0, "total_w_vat": 600.0},
        "pages": 1,
        "warnings": [],
    }
