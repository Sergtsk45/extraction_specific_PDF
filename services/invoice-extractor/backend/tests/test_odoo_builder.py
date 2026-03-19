"""Tests for app/odoo_builder.py and POST /convert?output=odoo_xlsx."""
import io
import os
import sys
import tempfile
from pathlib import Path

import pytest
import openpyxl

# Ensure backend on path (conftest already does this, but explicit is safer)
_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

from app.odoo_builder import build_odoo_xlsx, ODOO_HEADERS  # noqa: E402

MINIMAL_PDF = b"%PDF-1.4\n1 0 obj<</Type /Catalog>>endobj\n%%EOF\n"

SAMPLE_DATA = {
    "invoice_number": "INV-001",
    "supplier": {"inn": "1234567890", "name": "ООО Тест"},
    "items": [
        {
            "line_no": 1,
            "name": "Кабель ВВГнг 3x2.5",
            "article": "KBL-001",
            "unit": "м",
            "qty": 100,
            "price": 70.44,
            "vat_rate": "20",
            "vat_amount": 1408.8,
            "amount_w_vat": 8452.8,
            "discount": 0,
        }
    ],
}

MULTI_ITEM_DATA = {
    "invoice_number": "INV-002",
    "supplier": {"inn": "9876543210", "name": "ООО Другой"},
    "items": [
        {"line_no": 1, "name": "Товар А", "article": "", "unit": "шт", "qty": 5, "price": 100.0, "vat_rate": "20", "vat_amount": 100.0, "amount_w_vat": 600.0, "discount": 0},
        {"line_no": 2, "name": "Товар Б", "article": "X-99", "unit": "кг", "qty": 2, "price": 50.0, "vat_rate": "10", "vat_amount": 10.0, "amount_w_vat": 110.0, "discount": 0},
        {"line_no": 3, "name": "Товар В", "article": "", "unit": "", "qty": 1, "price": 200.0, "vat_rate": "", "vat_amount": 0, "amount_w_vat": 200.0, "discount": 0},
    ],
}


@pytest.fixture
def odoo_xlsx_path(tmp_path):
    return str(tmp_path / "test_odoo.xlsx")


# ── Unit tests: build_odoo_xlsx ────────────────────────────────────────────────

def test_build_odoo_xlsx_creates_file(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    assert Path(odoo_xlsx_path).exists()


def test_build_odoo_xlsx_sheet_name(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    wb = openpyxl.load_workbook(odoo_xlsx_path)
    assert "Template" in wb.sheetnames


def test_build_odoo_xlsx_headers_count(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    headers = [ws.cell(1, col).value for col in range(1, 10)]
    assert headers == ODOO_HEADERS


def test_build_odoo_xlsx_row_count(odoo_xlsx_path):
    build_odoo_xlsx(MULTI_ITEM_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    # строка 1 — заголовки, строки 2..N — данные
    data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in row)]
    assert len(data_rows) == 3


def test_build_odoo_xlsx_external_id_format(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    ext_id = ws.cell(2, 1).value
    assert ext_id == "inv_1234567890_INV-001_1"


def test_build_odoo_xlsx_name_field(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 2).value == "Кабель ВВГнг 3x2.5"


def test_build_odoo_xlsx_product_type_always_goods(odoo_xlsx_path):
    build_odoo_xlsx(MULTI_ITEM_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    for row in range(2, 5):
        assert ws.cell(row, 3).value == "Goods"


def test_build_odoo_xlsx_internal_reference(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 4).value == "KBL-001"


def test_build_odoo_xlsx_barcode_empty(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 5).value in (None, "")


def test_build_odoo_xlsx_sales_price_empty(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 6).value in (None, "")


def test_build_odoo_xlsx_cost_equals_price(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 7).value == 70.44


def test_build_odoo_xlsx_weight_empty(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 8).value in (None, "")


def test_build_odoo_xlsx_sales_description_contains_unit(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    desc = ws.cell(2, 9).value or ""
    assert "Ед.: м" in desc


def test_build_odoo_xlsx_sales_description_contains_vat(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    desc = ws.cell(2, 9).value or ""
    assert "НДС: 20%" in desc


def test_build_odoo_xlsx_sales_description_contains_qty(odoo_xlsx_path):
    build_odoo_xlsx(SAMPLE_DATA, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    desc = ws.cell(2, 9).value or ""
    assert "Кол-во: 100" in desc


def test_build_odoo_xlsx_zero_qty_no_error(odoo_xlsx_path):
    data = {
        "invoice_number": "X",
        "supplier": {"inn": "000"},
        "items": [{"line_no": 1, "name": "Test", "article": "", "unit": "", "qty": 0, "price": 50.0, "vat_rate": "", "vat_amount": 0, "amount_w_vat": 0, "discount": 0}],
    }
    build_odoo_xlsx(data, odoo_xlsx_path)  # не должно кинуть ZeroDivisionError
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 6).value in (None, "")  # Sales Price по-прежнему пустой


def test_build_odoo_xlsx_empty_items_only_headers(odoo_xlsx_path):
    build_odoo_xlsx({"invoice_number": "X", "supplier": {}, "items": []}, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.max_row == 1


def test_build_odoo_xlsx_missing_optional_fields_no_error(odoo_xlsx_path):
    data = {
        "invoice_number": "Y",
        "supplier": {},  # нет inn
        "items": [{"name": "Товар без полей"}],  # минимум полей
    }
    build_odoo_xlsx(data, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 1).value == "inv_x_Y_1"  # fallback inn="x"


def test_build_odoo_xlsx_inn_fallback_when_missing(odoo_xlsx_path):
    data = {"invoice_number": "Z", "items": [{"line_no": 5, "name": "A"}]}
    build_odoo_xlsx(data, odoo_xlsx_path)
    ws = openpyxl.load_workbook(odoo_xlsx_path)["Template"]
    assert ws.cell(2, 1).value == "inv_x_Z_5"


# ── Integration tests: POST /convert endpoint ──────────────────────────────────

def test_convert_odoo_xlsx_returns_200(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample_invoice(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    response = _upload(client, extra={"output": "odoo_xlsx"})
    assert response.status_code == 200


def test_convert_odoo_xlsx_content_type(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample_invoice(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    response = _upload(client, extra={"output": "odoo_xlsx"})
    assert "spreadsheetml" in response.content_type


def test_convert_odoo_xlsx_output_mode_header(client, monkeypatch, invoice_app_module):
    monkeypatch.setattr(invoice_app_module, "extract_invoice", lambda *a, **kw: (_sample_invoice(), False))
    monkeypatch.setattr(invoice_app_module, "validate_invoice_data", lambda d: [])
    response = _upload(client, extra={"output": "odoo_xlsx"})
    assert response.headers.get("X-Output-Mode") == "odoo_xlsx"


def test_convert_invalid_output_mode_returns_400(client):
    response = _upload(client, extra={"output": "invalid_mode"})
    assert response.status_code == 400


def test_convert_invalid_output_mode_has_error_field(client):
    response = _upload(client, extra={"output": "invalid_mode"})
    assert "error" in response.get_json()


# ── Helpers ────────────────────────────────────────────────────────────────────

def _upload(client, extra: dict | None = None):
    data = {"file": (io.BytesIO(MINIMAL_PDF), "invoice.pdf")}
    if extra:
        data.update(extra)
    return client.post("/convert", data=data, content_type="multipart/form-data")


def _sample_invoice() -> dict:
    return {
        "invoice_number": "TEST-001",
        "invoice_date": "2026-01-01",
        "supplier": {"inn": "1234567890", "name": "ООО Тест", "kpp": "", "address": "", "bank": {}},
        "buyer": {"inn": "", "name": "", "kpp": "", "address": ""},
        "items": [
            {"line_no": 1, "name": "Товар 1", "article": "A1", "unit": "шт", "qty": 5, "price": 100.0, "discount": 0, "vat_rate": "20", "vat_amount": 100.0, "amount_w_vat": 600.0},
        ],
        "totals": {"total_wo_vat": 500.0, "vat_total": 100.0, "total_w_vat": 600.0},
        "pages": 1,
        "warnings": [],
    }
