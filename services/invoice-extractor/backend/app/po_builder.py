"""
@file: po_builder.py
@description: Построитель XLSX для импорта заказа на закупку (purchase.order) в Odoo.
  Входные данные — invoice_data из extract_invoice().
  Формат совместим с Odoo Import: Закупки → Заказы → Импорт.
  Каждая строка товара = отдельная строка Excel; поля заголовка заказа
  заполняются только в первой строке (Odoo связывает строки по полю id).
@dependencies: openpyxl (уже в requirements.txt)
@created: 2026-04-20
"""

from __future__ import annotations
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

# Заголовки строго по именам полей модели purchase.order в Odoo.
# Odoo матчит по точному имени — пробел или регистр сломает импорт.
PO_HEADERS = [
    "id",                          # External ID заказа (связывает строки между собой)
    "name",                        # Номер заказа / ссылка на счёт
    "partner_id",                  # Поставщик (по отображаемому имени)
    "date_order",                  # Дата заказа
    "order_line/id",               # External ID строки заказа
    "order_line/product_id",       # Товар (по отображаемому имени)
    "order_line/name",             # Описание строки
    "order_line/product_qty",      # Количество
    "order_line/price_unit",       # Цена за единицу без НДС
    "order_line/product_uom",      # Единица измерения
    "order_line/taxes_id",         # Налог (например: "20.0%")
]

# Стили заголовка
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _num(value, default: float = 0.0) -> float:
    """Число из позиции счёта (LLM может отдать str с запятой)."""
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        s = str(value).strip().replace(" ", "").replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return default


def _tax_display_for_odoo(vat_rate) -> str:
    """Строка для order_line/taxes_id; пусто если ставку распознать нельзя."""
    if vat_rate is None:
        return ""
    if isinstance(vat_rate, str):
        s = vat_rate.strip()
        if not s:
            return ""
        s = s.rstrip("%").strip()
        if not s:
            return ""
        try:
            rate = float(s.replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            return ""
    else:
        try:
            rate = float(vat_rate)
        except (ValueError, TypeError):
            return ""
    return f"{rate:.10g}%"


def build_po_xlsx(data: dict, output_path: str) -> None:
    """Строит XLSX для импорта заказа на закупку в Odoo (purchase.order).

    Структура файла:
      - Строка 1: заголовки (PO_HEADERS)
      - Строка 2: первая позиция счёта + поля заголовка заказа
      - Строки 3+: остальные позиции; поля заголовка заказа пустые,
        только id заказа остаётся для связи

    Args:
        data: Словарь invoice_data от extract_invoice().
        output_path: Путь для сохранения .xlsx файла.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "purchase.order"

    # ── Заголовки с форматированием ──────────────────────────────────────────
    ws.append(PO_HEADERS)
    for col_idx, _ in enumerate(PO_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 18

    # ── Извлечение данных из invoice_data ────────────────────────────────────
    invoice_num = data.get("invoice_number") or "unknown"
    invoice_date = data.get("date") or data.get("invoice_date") or ""
    supplier = data.get("supplier") or {}
    inn = supplier.get("inn") or "x"
    supplier_name = (supplier.get("name") or "").strip()
    items = data.get("items") or []

    # External ID заказа: детерминированный, повторный импорт обновляет, не дублирует
    order_ext_id = f"po_{inn}_{invoice_num}"

    # ── Строки товаров ───────────────────────────────────────────────────────
    for idx, item in enumerate(items, start=1):
        line_no = item.get("line_no") or idx
        qty = _num(item.get("qty"))
        price = _num(item.get("price"))  # цена без НДС
        product_name = (item.get("name") or "").strip()
        unit = (item.get("unit") or "").strip()

        tax_str = _tax_display_for_odoo(item.get("vat_rate"))

        line_ext_id = f"{order_ext_id}_line_{line_no}"

        # Поля заголовка заказа: только в первой строке (idx==1).
        # Для строк 2+ оставляем только id заказа — Odoo сам связывает строки.
        order_name = invoice_num if idx == 1 else ""
        order_partner = supplier_name if idx == 1 else ""
        order_date = invoice_date if idx == 1 else ""

        ws.append([
            order_ext_id,    # id
            order_name,      # name
            order_partner,   # partner_id
            order_date,      # date_order
            line_ext_id,     # order_line/id
            product_name,    # order_line/product_id
            product_name,    # order_line/name (описание = название товара)
            qty,             # order_line/product_qty
            price,           # order_line/price_unit
            unit,            # order_line/product_uom
            tax_str,         # order_line/taxes_id
        ])

    # ── Ширина колонок для читаемости ────────────────────────────────────────
    col_widths = [28, 22, 30, 14, 32, 40, 40, 12, 14, 10, 10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    wb.save(output_path)
    logger.info(f"PO XLSX saved: {output_path}, order_id: {order_ext_id}, lines: {len(items)}")
