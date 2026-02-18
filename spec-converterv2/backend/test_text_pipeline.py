#!/usr/bin/env python3
"""
Тестовый скрипт для проверки text-first pipeline.
"""

import sys
import os

# Добавляем путь к модулям
sys.path.insert(0, os.path.dirname(__file__))

from pdf_text_extractor import has_text_layer, extract_table_from_text, normalize_table_to_9cols
from spec_utils import SpecificationExcelBuilder
import pdfplumber

def test_text_pipeline():
    """Тестирование text-first pipeline на эталонном PDF"""
    
    pdf_path = "../Образец/Спецификация 2273-1.1-ВК2 Офисы.pdf"
    output_path = "../output/test_text_pipeline_result.xlsx"
    
    print("="*70)
    print("ТЕСТ: Text-first PDF Pipeline")
    print("="*70)
    print(f"\nPDF: {pdf_path}")
    
    # Шаг 1: Проверка текстового слоя
    print("\n1. Проверка текстового слоя...")
    text_flags = has_text_layer(pdf_path)
    print(f"   Результат: {text_flags}")
    print(f"   Страниц с текстом: {sum(text_flags)}/{len(text_flags)}")
    
    # Шаг 2: Извлечение данных
    print("\n2. Извлечение данных...")
    pages_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            page_num = i + 1
            print(f"\n   Страница {page_num}:")
            
            if text_flags[i]:
                print(f"      Режим: ТЕКСТ")
                try:
                    raw_rows, pdf_header = extract_table_from_text(page)
                    print(f"      Извлечено строк: {len(raw_rows)}")
                    
                    if raw_rows and len(raw_rows) > 1:
                        page_data = normalize_table_to_9cols(raw_rows, page_num, pdf_header)
                        page_data['page_num'] = page_num
                        pages_data.append(page_data)
                        print(f"      ✓ Лист: '{page_data['sheet_name']}'")
                        print(f"      ✓ Строк после нормализации: {len(page_data['rows'])}")
                        
                        # Показываем первые 3 строки данных
                        print(f"      Первые строки:")
                        for idx, row in enumerate(page_data['rows'][:3]):
                            row_preview = [str(c)[:30] for c in row[:3]]  # первые 3 колонки
                            print(f"         [{idx}]: {row_preview}")
                    else:
                        print(f"      ⚠️ Пустая таблица")
                except Exception as e:
                    print(f"      ❌ Ошибка: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"      Режим: VISION (пропущено в тесте)")
    
    # Шаг 3: Создание Excel
    print(f"\n3. Создание Excel...")
    print(f"   Листов для создания: {len(pages_data)}")
    
    if pages_data:
        os.makedirs("../output", exist_ok=True)
        builder = SpecificationExcelBuilder(output_path)
        
        for page_data in pages_data:
            rows = page_data.get('rows', [])
            sheet_name = page_data.get('sheet_name', 'Спецификация')
            
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + "..."
            
            print(f"   Создание листа '{sheet_name}': {len(rows)} строк")
            builder.create_sheet(sheet_name, rows)
        
        builder.save()
        print(f"\n✅ Excel сохранён: {output_path}")
        print(f"   Размер: {os.path.getsize(output_path)} байт")
    else:
        print(f"\n⚠️ Нет данных для Excel")
    
    # Шаг 4: Сравнение с эталоном
    print(f"\n4. Сравнение с эталоном...")
    etalon_path = "../Образец/Спецификация Эталон из web_cloude_ВК2_Офисы.xlsx"
    
    if os.path.exists(etalon_path):
        import openpyxl
        
        # Читаем эталон
        etalon_wb = openpyxl.load_workbook(etalon_path)
        print(f"   Эталон: {len(etalon_wb.sheetnames)} листов")
        for name in etalon_wb.sheetnames:
            sheet = etalon_wb[name]
            print(f"      - '{name}': {sheet.max_row} строк")
        
        # Читаем результат
        if os.path.exists(output_path):
            result_wb = openpyxl.load_workbook(output_path)
            print(f"   Результат: {len(result_wb.sheetnames)} листов")
            for name in result_wb.sheetnames:
                sheet = result_wb[name]
                print(f"      - '{name}': {sheet.max_row} строк")
            
            # Сравнение количества листов
            if len(etalon_wb.sheetnames) == len(result_wb.sheetnames):
                print(f"   ✓ Количество листов совпадает: {len(etalon_wb.sheetnames)}")
            else:
                print(f"   ✗ Количество листов НЕ совпадает: эталон={len(etalon_wb.sheetnames)}, результат={len(result_wb.sheetnames)}")
    else:
        print(f"   ⚠️ Эталон не найден: {etalon_path}")
    
    print("\n" + "="*70)
    print("ТЕСТ ЗАВЕРШЁН")
    print("="*70)

if __name__ == "__main__":
    test_text_pipeline()
